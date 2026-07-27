"""One-time migration from the startup Excel register into the active SQL DB."""

import os
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import select

from customer_store import CustomerRecord, SQLCustomerRepository

load_dotenv(Path(__file__).with_name(".env"))


def migrate() -> None:
    excel_path = Path(os.getenv("CUSTOMER_EXCEL_PATH", "data/customers.xlsx"))
    if not excel_path.exists():
        print(f"No Excel register found at {excel_path}; nothing to migrate.")
        return

    repository = SQLCustomerRepository(os.environ["DATABASE_URL"])
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook["Customers"]
    inserted = 0
    skipped = 0

    with repository.session_factory.begin() as session:
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values[0]:
                continue
            customer_id, created_at, name, company, email, phone, requirement, source, status = values
            # Idempotent migration: running the script again does not duplicate leads.
            if session.scalar(select(CustomerRecord.customer_id).where(CustomerRecord.customer_id == customer_id)):
                skipped += 1
                continue
            session.add(CustomerRecord(
                customer_id=customer_id,
                created_at=created_at,
                name=name,
                company=company or "",
                email=email,
                phone=phone or "",
                requirement=requirement,
                source=source or "Excel migration",
                status=status or "New",
            ))
            inserted += 1

    print(f"Migration complete: {inserted} inserted, {skipped} already present.")


if __name__ == "__main__":
    migrate()
