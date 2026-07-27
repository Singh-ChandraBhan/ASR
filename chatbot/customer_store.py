import os
import threading
import logging
from abc import ABC, abstractmethod
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import DateTime, String, Text, create_engine
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, sessionmaker

logger = logging.getLogger(__name__)


class CustomerRepository(ABC):
    """Common storage contract used by both Excel and SQL backends.

    REMARK: The API depends only on this interface. This keeps the website and
    endpoint unchanged when the company migrates from Excel to PostgreSQL.
    """

    @abstractmethod
    def create(self, customer: dict, customer_id: str | None = None) -> str:
        raise NotImplementedError


class ExcelCustomerRepository(CustomerRepository):
    """Startup-stage storage that appends customer enquiries to one workbook.

    REMARK: Excel is appropriate for low traffic and a single API process. It
    should not be used by multiple servers because an .xlsx file is not a
    transactional database.
    """
    headers = ["Customer ID", "Created At (UTC)", "Name", "Company", "Email", "Phone", "Requirement", "Source", "Status"]

    def __init__(self, path: str):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        # Prevent two requests in this API process from writing simultaneously.
        self._lock = threading.Lock()
        if not self.path.exists():
            self._create_workbook()

    def _create_workbook(self):
        # Create a usable register automatically if the configured file is absent.
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        sheet.append(self.headers)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="0B2949")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = "A1:I1"
        for index, width in enumerate([22, 23, 24, 24, 30, 18, 55, 16, 16], start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        workbook.save(self.path)

    def create(self, customer: dict, customer_id: str | None = None) -> str:
        # Public IDs allow staff to reference an enquiry without exposing row IDs.
        customer_id = customer_id or f"ASR-{uuid4().hex[:10].upper()}"
        row = [customer_id, datetime.now(timezone.utc).replace(microsecond=0).isoformat(), customer["name"], customer.get("company", ""), customer["email"], customer.get("phone", ""), customer["requirement"], customer.get("source", "Website"), "New"]
        # Keep the read/append/save sequence inside one critical section.
        with self._lock:
            workbook = load_workbook(self.path)
            sheet = workbook["Customers"]
            sheet.append(row)
            last_row = sheet.max_row
            sheet.cell(last_row, 7).alignment = Alignment(wrap_text=True, vertical="top")
            existing = sheet.tables.get("CustomersTable")
            if existing:
                existing.ref = f"A1:I{last_row}"
            else:
                table = Table(displayName="CustomersTable", ref=f"A1:I{last_row}")
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                sheet.add_table(table)
            workbook.save(self.path)
        return customer_id


class Base(DeclarativeBase):
    pass


class CustomerRecord(Base):
    """Production customer-lead model mapped to the SQL table.

    REMARK: Lengths match API validation so invalid or oversized values are
    rejected before reaching the database.
    """

    __tablename__ = "customer_leads"

    customer_id: Mapped[str] = mapped_column(String(32), primary_key=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False, index=True)
    name: Mapped[str] = mapped_column(String(100), nullable=False)
    company: Mapped[str] = mapped_column(String(150), nullable=False, default="")
    email: Mapped[str] = mapped_column(String(254), nullable=False, index=True)
    phone: Mapped[str] = mapped_column(String(30), nullable=False, default="")
    requirement: Mapped[str] = mapped_column(Text, nullable=False)
    source: Mapped[str] = mapped_column(String(50), nullable=False, default="Website")
    status: Mapped[str] = mapped_column(String(30), nullable=False, default="New", index=True)


class SQLCustomerRepository(CustomerRepository):
    """Transactional SQL storage for higher traffic and multiple API servers."""

    def __init__(self, database_url: str):
        if not database_url:
            raise RuntimeError("DATABASE_URL is required when CUSTOMER_STORAGE=sql.")
        # pool_pre_ping replaces stale database connections before a request uses them.
        self.engine = create_engine(database_url, pool_pre_ping=True)
        self.session_factory = sessionmaker(bind=self.engine, expire_on_commit=False)
        # Safe for initial startup; use Alembic migrations once the schema evolves.
        Base.metadata.create_all(self.engine)

    def create(self, customer: dict, customer_id: str | None = None) -> str:
        customer_id = customer_id or f"ASR-{uuid4().hex[:10].upper()}"
        record = CustomerRecord(
            customer_id=customer_id,
            created_at=datetime.now(timezone.utc),
            name=customer["name"],
            company=customer.get("company", ""),
            email=customer["email"],
            phone=customer.get("phone", ""),
            requirement=customer["requirement"],
            source=customer.get("source", "Website"),
            status="New",
        )
        # The context manager commits on success and rolls back on any exception.
        with self.session_factory.begin() as session:
            session.add(record)
        return customer_id


class DualCustomerRepository(CustomerRepository):
    """Save to SQL first, then mirror the same lead into Excel.

    REMARK: SQL is the source of truth. Excel is a convenient operational copy,
    so an Excel lock must not cause a successfully committed SQL lead to fail.
    """

    def __init__(self, sql_repository: SQLCustomerRepository, excel_repository: ExcelCustomerRepository):
        self.sql_repository = sql_repository
        self.excel_repository = excel_repository

    def create(self, customer: dict, customer_id: str | None = None) -> str:
        shared_id = customer_id or f"ASR-{uuid4().hex[:10].upper()}"
        self.sql_repository.create(customer, shared_id)
        try:
            self.excel_repository.create(customer, shared_id)
        except Exception:
            logger.exception("Customer %s is safe in SQL, but the Excel mirror failed.", shared_id)
        return shared_id


@lru_cache
def get_customer_repository() -> CustomerRepository:
    """Build the selected repository once and reuse its connection/lock.

    REMARK: CUSTOMER_STORAGE=excel is the safe startup default. Change it to
    `sql` only after DATABASE_URL has been configured and secured.
    """
    backend = os.getenv("CUSTOMER_STORAGE", "excel").lower()
    if backend == "excel":
        default_path = Path(__file__).with_name("data") / "customers.xlsx"
        return ExcelCustomerRepository(os.getenv("CUSTOMER_EXCEL_PATH", str(default_path)))
    if backend == "sql":
        return SQLCustomerRepository(os.getenv("DATABASE_URL", ""))
    if backend == "both":
        default_path = Path(__file__).with_name("data") / "customers.xlsx"
        return DualCustomerRepository(
            SQLCustomerRepository(os.getenv("DATABASE_URL", "")),
            ExcelCustomerRepository(os.getenv("CUSTOMER_EXCEL_PATH", str(default_path))),
        )
    raise RuntimeError("CUSTOMER_STORAGE must be 'excel', 'sql', or 'both'.")
